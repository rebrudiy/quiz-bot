"""
quiz_bot.py — простой Telegram-бот викторина.

Зависимости:
    pip install pytelegrambotapi openpyxl

Запуск:
    python quiz_bot.py

Настройки лежат в config.py
"""

import random
from typing import List, Dict, Any

import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton
import openpyxl

import config


# ---------- Загрузка вопросов из Excel ----------

def load_questions_from_xls(xls_path: str) -> List[Dict[str, Any]]:
    """
    Читает вопросы из Excel и возвращает список.

    Формат таблицы (1-я строка заголовки):
        question | option1 | option2 | option3 | option4 | correct_option (1-4)

    Возвращает список:
    {
        "text": "...",
        "options": ["...", "...", "...", "..."],
        "correct_index": 0..3
    }
    """
    wb = openpyxl.load_workbook(xls_path)
    sheet = wb.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {name: idx for idx, name in enumerate(header_row)}

    required_columns = [
        "question", "option1", "option2", "option3", "option4", "correct_option"
    ]
    for col in required_columns:
        if col not in headers:
            raise ValueError(f"Нет столбца '{col}' в Excel")

    questions = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue

        q_text = row[headers["question"]]
        o1 = row[headers["option1"]]
        o2 = row[headers["option2"]]
        o3 = row[headers["option3"]]
        o4 = row[headers["option4"]]
        correct_raw = row[headers["correct_option"]]

        # если что-то не заполнено — пропускаем
        if not (q_text and o1 and o2 and o3 and o4):
            continue

        try:
            correct_num = int(correct_raw)
        except (TypeError, ValueError):
            continue

        if not (1 <= correct_num <= 4):
            continue

        questions.append({
            "text": str(q_text),
            "options": [str(o1), str(o2), str(o3), str(o4)],
            "correct_index": correct_num - 1
        })

    return questions


# ---------- Состояния пользователей ----------

class QuizState:
    """Состояние викторины для одного пользователя."""

    def __init__(self, questions: List[Dict[str, Any]]):
        self.questions = questions
        self.current_index = 0
        self.score = 0
        self.active = False

    def current_question(self) -> Dict[str, Any]:
        if 0 <= self.current_index < len(self.questions):
            return self.questions[self.current_index]
        return None

    def next(self):
        self.current_index += 1

    def finished(self) -> bool:
        return self.current_index >= len(self.questions)


# ---------- Главная функция ----------

def main():
    questions = load_questions_from_xls(config.XLS_PATH)

    if not questions:
        print("Вопросов не найдено. Проверь XLS_PATH и формат Excel.")
        return

    if config.SHUFFLE_QUESTIONS:
        random.shuffle(questions)

    bot = telebot.TeleBot(config.TOKEN)
    states: Dict[int, QuizState] = {}

    def get_state(user_id: int) -> QuizState:
        if user_id not in states:
            states[user_id] = QuizState(questions)
        return states[user_id]

    def send_question(chat_id: int, state: QuizState):
        q = state.current_question()
        if not q:
            bot.send_message(chat_id, "Вопросов нет.")
            return

        kb = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        for opt in q["options"]:
            kb.add(KeyboardButton(opt))

        text = f"Вопрос {state.current_index + 1}/{len(state.questions)}:\n\n{q['text']}"
        bot.send_message(chat_id, text, reply_markup=kb)

    # --- команды ---

    @bot.message_handler(commands=["start"])
    def start_cmd(message):
        state = get_state(message.from_user.id)
        state.active = False
        state.current_index = 0
        state.score = 0

        bot.send_message(
            message.chat.id,
            "Привет! Это бот-викторина.\n"
            "Нажми /quiz чтобы начать."
        )

    @bot.message_handler(commands=["quiz"])
    def quiz_cmd(message):
        state = get_state(message.from_user.id)

        state.current_index = 0
        state.score = 0
        state.active = True

        bot.send_message(message.chat.id, "Поехали! 🚀")
        send_question(message.chat.id, state)

    # --- ответы ---

    @bot.message_handler(func=lambda m: True, content_types=["text"])
    def any_text(message):
        user_id = message.from_user.id
        state = get_state(user_id)

        if not state.active:
            if not message.text.startswith("/"):
                bot.send_message(message.chat.id, "Нажми /quiz чтобы начать викторину.")
            return

        q = state.current_question()
        if not q:
            bot.send_message(message.chat.id, "Вопросы закончились. Нажми /quiz заново.")
            state.active = False
            return

        answer = message.text.strip()
        correct = q["options"][q["correct_index"]]

        if answer == correct:
            state.score += 1
            bot.send_message(message.chat.id, "✅ Правильно!")
        else:
            if config.SHOW_CORRECT_ON_WRONG:
                bot.send_message(message.chat.id, f"❌ Неправильно. Правильный ответ: {correct}")
            else:
                bot.send_message(message.chat.id, "❌ Неправильно.")

        state.next()

        if state.finished():
            bot.send_message(
                message.chat.id,
                f"Конец викторины!\nРезультат: {state.score}/{len(state.questions)}"
            )
            state.active = False
        else:
            send_question(message.chat.id, state)

    print("Бот запущен...")
    bot.infinity_polling()


if __name__ == "__main__":
    main()
